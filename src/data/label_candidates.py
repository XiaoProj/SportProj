"""Build small candidate label tables from confirmed spreadsheet fields."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any


MENTAL_FATIGUE_COLUMNS = {
    "shot_1_accuracy": 8,
    "shot_1_accuracy_1_20": 9,
    "shot_1_accuracy_21_40": 10,
    "shot_1_accuracy_41_60": 11,
    "shot_2_accuracy": 12,
    "shot_2_accuracy_1_20": 13,
    "shot_2_accuracy_21_40": 14,
    "shot_2_accuracy_41_60": 15,
    "prontezza": 16,
    "vas_mf_1": 17,
    "vas_mf_2": 18,
    "vas_mf_3": 19,
    "vas_mot_1": 20,
    "vas_mot_2": 21,
    "vas_mot_3": 22,
    "questions_accuracy": 23,
}

SLEEP_DEPRIVATION_COLUMNS = {
    "shot_1_accuracy": 24,
    "shot_1_accuracy_1_20": 25,
    "shot_1_accuracy_21_40": 26,
    "shot_1_accuracy_41_60": 27,
    "shot_2_accuracy": 28,
    "shot_2_accuracy_1_20": 29,
    "shot_2_accuracy_21_40": 30,
    "shot_2_accuracy_41_60": 31,
    "prontezza": 32,
    "vas_mf_1": 33,
    "vas_mf_2": 34,
    "vas_mf_3": 35,
    "vas_mot_1": 36,
    "vas_mot_2": 37,
    "vas_mot_3": 38,
    "questions_accuracy": 39,
}

LABEL_CANDIDATE_COLUMNS = [
    "dataset_name",
    "subject_id",
    "condition_group",
    "fatigue_condition_candidate",
    "sleep_condition_candidate",
    "before_after_label_candidate",
    "fatigue_label_candidate",
    "fatigue_label_candidate_rule",
    "shot_1_accuracy",
    "shot_1_accuracy_1_20",
    "shot_1_accuracy_21_40",
    "shot_1_accuracy_41_60",
    "shot_2_accuracy",
    "shot_2_accuracy_1_20",
    "shot_2_accuracy_21_40",
    "shot_2_accuracy_41_60",
    "prontezza",
    "vas_mf_1",
    "vas_mf_2",
    "vas_mf_3",
    "vas_mot_1",
    "vas_mot_2",
    "vas_mot_3",
    "questions_accuracy",
]


@dataclass(frozen=True)
class LabelCandidateSummary:
    """Summary for the generated label-candidate table."""

    output_path: Path
    rows: int
    columns: int
    fatigue_sleep_fields: list[str]
    performance_fields: list[str]
    has_explicit_fatigue_label: bool
    has_explicit_before_after_label: bool


def _cell(row: tuple[Any, ...], one_based_index: int) -> Any:
    zero_based = one_based_index - 1
    if zero_based >= len(row):
        return None
    return row[zero_based]


def _build_condition_row(subject_id: Any, condition_group: str, row: tuple[Any, ...], mapping: dict[str, int]) -> dict[str, Any]:
    output = {
        "dataset_name": "basketball_fatigue_sleep_5742821",
        "subject_id": subject_id,
        "condition_group": condition_group,
        "fatigue_condition_candidate": "mental_fatigue_condition_present",
        "sleep_condition_candidate": (
            "sleep_deprivation_condition_present"
            if condition_group == "mental_fatigue_plus_sleep_deprivation"
            else "no_sleep_deprivation_condition_label_observed"
        ),
        "before_after_label_candidate": "",
        "fatigue_label_candidate": "",
        "fatigue_label_candidate_rule": "candidate only; no explicit binary fatigue_label column observed",
    }
    for field, index in mapping.items():
        output[field] = _cell(row, index)
    return output


def build_fatigue_sleep_label_candidates(xlsx_path: Path, output_path: Path) -> LabelCandidateSummary:
    """Build a long-form candidate table from Data.xlsx without modifying it."""

    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to read Data.xlsx metadata.") from exc

    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    candidate_rows: list[dict[str, Any]] = []
    for row in rows[2:]:
        subject_id = _cell(row, 2)
        if subject_id is None:
            continue
        candidate_rows.append(
            _build_condition_row(
                subject_id,
                "mental_fatigue",
                row,
                MENTAL_FATIGUE_COLUMNS,
            )
        )
        candidate_rows.append(
            _build_condition_row(
                subject_id,
                "mental_fatigue_plus_sleep_deprivation",
                row,
                SLEEP_DEPRIVATION_COLUMNS,
            )
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=LABEL_CANDIDATE_COLUMNS)
        writer.writeheader()
        for row in candidate_rows:
            writer.writerow({column: row.get(column) for column in LABEL_CANDIDATE_COLUMNS})

    return LabelCandidateSummary(
        output_path=output_path,
        rows=len(candidate_rows),
        columns=len(LABEL_CANDIDATE_COLUMNS),
        fatigue_sleep_fields=[
            "condition_group",
            "fatigue_condition_candidate",
            "sleep_condition_candidate",
            "vas_mf_1",
            "vas_mf_2",
            "vas_mf_3",
            "vas_mot_1",
            "vas_mot_2",
            "vas_mot_3",
        ],
        performance_fields=[
            "shot_1_accuracy",
            "shot_1_accuracy_1_20",
            "shot_1_accuracy_21_40",
            "shot_1_accuracy_41_60",
            "shot_2_accuracy",
            "shot_2_accuracy_1_20",
            "shot_2_accuracy_21_40",
            "shot_2_accuracy_41_60",
            "questions_accuracy",
        ],
        has_explicit_fatigue_label=False,
        has_explicit_before_after_label=False,
    )
