"""Small-sample schema probing helpers.

The functions in this module read selected metadata/sample files only. They do
not write into raw data directories and do not perform extraction or conversion.
"""

from __future__ import annotations

import json
import os
from json import JSONDecodeError
from pathlib import Path
from typing import Any


def safe_relpath(path: Path, root: Path) -> str:
    """Return a portable relative path when possible."""

    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except ValueError:
        return path.as_posix()


def scalar_preview(value: Any) -> Any:
    """Convert scalar values into JSON-safe short previews."""

    if value is None or isinstance(value, (bool, int, float)):
        return value
    text = str(value)
    return text if len(text) <= 120 else text[:117] + "..."


def summarize_value(value: Any, depth: int = 0, max_depth: int = 3, max_keys: int = 10) -> dict[str, Any]:
    """Summarize nested JSON-like data without storing full content."""

    if isinstance(value, dict):
        keys = list(value.keys())
        summary: dict[str, Any] = {
            "type": "dict",
            "length": len(value),
            "keys": [str(key) for key in keys[:max_keys]],
        }
        if depth < max_depth:
            summary["sample"] = {
                str(key): summarize_value(value[key], depth + 1, max_depth, max_keys)
                for key in keys[: min(3, len(keys))]
            }
        return summary
    if isinstance(value, list):
        summary = {"type": "list", "length": len(value)}
        if value and depth < max_depth:
            summary["first"] = summarize_value(value[0], depth + 1, max_depth, max_keys)
        return summary
    return {"type": type(value).__name__, "value": scalar_preview(value)}


def summarize_json_file(path: Path, project_root: Path, max_depth: int = 3) -> dict[str, Any]:
    """Load one selected JSON file and summarize its structure."""

    try:
        with path.open("r", encoding="utf-8") as handle:
            data = json.load(handle)
    except JSONDecodeError as exc:
        excerpt = read_text_excerpt(path, project_root, max_chars=800)
        return {
            "path": safe_relpath(path, project_root),
            "size_bytes": path.stat().st_size,
            "parse_status": "json_decode_error",
            "error": str(exc),
            "excerpt": excerpt["excerpt"],
        }
    return {
        "path": safe_relpath(path, project_root),
        "size_bytes": path.stat().st_size,
        "parse_status": "ok",
        "summary": summarize_value(data, max_depth=max_depth),
    }


def read_text_excerpt(path: Path, project_root: Path, max_chars: int = 2000) -> dict[str, Any]:
    """Read a short text excerpt from a selected documentation file."""

    data = path.read_bytes()[: max_chars * 4]
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            text = data.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = data.decode("utf-8", errors="replace")
    return {
        "path": safe_relpath(path, project_root),
        "size_bytes": path.stat().st_size,
        "excerpt": text[:max_chars],
    }


def first_matching_file(root: Path, directory_name: str, suffix: str = ".json") -> Path | None:
    """Find one sample file below a named directory, stopping after the first hit."""

    for dirpath, _dirnames, filenames in os.walk(root):
        current = Path(dirpath)
        if current.name != directory_name:
            continue
        for filename in sorted(filenames):
            if filename.lower().endswith(suffix.lower()):
                return current / filename
    return None


def list_immediate_files(root: Path, pattern: str, project_root: Path | None = None) -> list[dict[str, Any]]:
    """List file names and sizes in one selected directory without reading content."""

    if not root.is_dir():
        return []
    return [
        {
            "name": path.name,
            "path": safe_relpath(path, project_root) if project_root is not None else path.as_posix(),
            "size_bytes": path.stat().st_size,
        }
        for path in sorted(root.glob(pattern))
        if path.is_file()
    ]


def read_xlsx_preview(path: Path, project_root: Path, max_rows: int = 6) -> dict[str, Any]:
    """Read workbook sheet names and the first rows with openpyxl if available."""

    try:
        import openpyxl
    except ImportError:
        return {
            "path": safe_relpath(path, project_root),
            "dependency": "openpyxl_missing",
            "message": "Install openpyxl to read Excel workbook metadata.",
        }

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: list[dict[str, Any]] = []
        for worksheet in workbook.worksheets:
            rows: list[list[Any]] = []
            for idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                if idx >= max_rows:
                    break
                rows.append([scalar_preview(value) for value in row])
            sheets.append(
                {
                    "name": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "preview_rows": rows,
                }
            )
        return {
            "path": safe_relpath(path, project_root),
            "size_bytes": path.stat().st_size,
            "dependency": "openpyxl_available",
            "sheets": sheets,
        }
    finally:
        workbook.close()
